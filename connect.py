import os
import openpyxl


symbols = ["A", "B", "C", "D", "E", "F", "G", "H", 'I', 'J', "K", 'L', 'M', 'N', 'O', 'P', 'Q', "R", 'S', 'T', 'U', 'V',
           'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO',
           'AP', 'AQ', 'AR', 'AS', 'AT', 'AU', 'AV', 'AW', 'AX', 'AY', 'AZ', 'BA', 'BB', 'BC', 'BD', 'BE', 'BF', 'BG',
           'BH', 'BI', 'BJ', 'BK', 'BL', 'BM', 'BN', 'BO', 'BP', 'BQ', 'BR', 'BS', 'BT', 'BU', 'BV', 'BW', 'BX', 'BY',
           'BZ']


def get_unfinished_products():
    workbook = openpyxl.load_workbook("Входящие данные.xlsx")
    page = workbook.active
    page["G2"].value = "Товар обработан"
    result = list()
    for index in range(5, page.max_row + 1):
        is_finished = page[f"G{index}"].value
        if is_finished != "Да":
            sub_result = {"index": index, "product_id": page[f"E{index}"].value}
            result.append(sub_result)
    workbook.save("Входящие данные.xlsx")
    return result


def set_unfinished_all_products():
    workbook = openpyxl.load_workbook("Входящие данные.xlsx")
    page = workbook.active
    for index in range(5, page.max_row + 1):
        page[f"G{index}"].value = ""
    workbook.save("Входящие данные.xlsx")


def set_finished_product(index):
    workbook = openpyxl.load_workbook("Входящие данные.xlsx")
    page = workbook.active
    page[f"G{index}"].value = "Да"
    workbook.save("Входящие данные.xlsx")


def get_finally_symbol(page):
    for symbol in symbols:
        if page[f"{symbol}1"].value is None:
            return symbol


def record_characteristics(file_name, characteristics, index):
    workbook = openpyxl.load_workbook(f"result/{file_name}")
    page = workbook.active
    for characteristic in characteristics:
        field_exist = False
        for symbol in symbols:
            current_name = page[f"{symbol}1"].value
            if page[f"{symbol}1"].value is None:
                break
            if characteristic["name"] == current_name:
                page[f"{symbol}{index}"].value = characteristic["value"]
                field_exist = True
                break
        if not field_exist:
            last_symbol = get_finally_symbol(page=page)
            page[f"{last_symbol}1"].value = characteristic["name"]
            page[f"{last_symbol}{index}"].value = characteristic["value"]
    workbook.save(f"result/{file_name}")


def record_product_data(file_name, product_id, title, description, product_link, product_code, price, currency,
                        category_one, category_two, category_three, category_four, images):
    if file_name in os.listdir("result"):
        workbook = openpyxl.load_workbook(f"result/{file_name}")
        page = workbook.active
        current_index = page.max_row + 1
        page[f"A{current_index}"].value = category_one
        page[f"B{current_index}"].value = category_two
        page[f"C{current_index}"].value = category_three
        page[f"D{current_index}"].value = category_four
        page[f"E{current_index}"].value = product_id
        page[f"F{current_index}"].value = title
        page[f"G{current_index}"].value = price
        page[f"H{current_index}"].value = currency
        page[f"I{current_index}"].value = "В наличии"
        page[f"J{current_index}"].value = description
        page[f"K{current_index}"].value = images
        page[f"L{current_index}"].value = product_link
        page[f"M{current_index}"].value = product_code
    else:
        current_index = 1
        workbook = openpyxl.Workbook()
        page = workbook.active
        page[f"A{current_index}"].value = "Категория 1"
        page[f"B{current_index}"].value = "Категория 2"
        page[f"C{current_index}"].value = "Категория 3"
        page[f"D{current_index}"].value = "Категория 4"
        page[f"E{current_index}"].value = "Артикул"
        page[f"F{current_index}"].value = "Название"
        page[f"G{current_index}"].value = "Цена"
        page[f"H{current_index}"].value = "Валюта"
        page[f"I{current_index}"].value = "Наличие"
        page[f"J{current_index}"].value = "Описание"
        page[f"K{current_index}"].value = "Изображения"
        page[f"L{current_index}"].value = "Ссылка на товар"
        page[f"M{current_index}"].value = "Код товара"
        current_index += 1
        page[f"A{current_index}"].value = category_one
        page[f"B{current_index}"].value = category_two
        page[f"C{current_index}"].value = category_three
        page[f"D{current_index}"].value = category_four
        page[f"E{current_index}"].value = product_id
        page[f"F{current_index}"].value = title
        page[f"G{current_index}"].value = price
        page[f"H{current_index}"].value = currency
        page[f"I{current_index}"].value = "В наличии"
        page[f"J{current_index}"].value = description
        page[f"K{current_index}"].value = images
        page[f"L{current_index}"].value = product_link
        page[f"M{current_index}"].value = product_code
    workbook.save(f"result/{file_name}")
    return current_index


if __name__ == "__main__":
    get_unfinished_products()
